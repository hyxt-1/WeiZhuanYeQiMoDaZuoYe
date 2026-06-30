from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from werkzeug.security import check_password_hash, generate_password_hash
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///movie_history.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# 数据库模型
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    user_name = db.Column(db.String(100), nullable=False)
    password = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    records = db.relationship('MovieRecord', backref='user', lazy=True)

class MovieRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    type = db.Column(db.String(50), nullable=False)  # 电影、电视剧、综艺等
    watch_time = db.Column(db.DateTime, nullable=True)
    location = db.Column(db.String(100))
    people = db.Column(db.String(200))
    event = db.Column(db.Text)
    rating = db.Column(db.Integer)  # 评分 1-5
    comment = db.Column(db.Text)
    tags = db.Column(db.String(200))  # 标签，用逗号分隔
    is_shared = db.Column(db.Boolean, default=False)  # 是否分享
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# 创建数据库
with app.app_context():
    db.create_all()

# 路由
@app.route('/')
def index():
    if 'user_id' in session:
        return render_template('index.html', 
                               logged_in=True, 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']))
    return render_template('index.html', logged_in=False)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form.get('confirmPassword')
        
        # 验证密码是否匹配
        if password != confirm_password:
            return jsonify({'error': '两次输入的密码不一致，请重新输入'}), 400
        
        # 检查用户名是否已存在
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            return jsonify({'error': '用户名已存在'}), 400
        
        # 创建新用户，默认用户名称与用户名相同
        new_user = User(username=username, user_name=username, password=generate_password_hash(password))
        db.session.add(new_user)
        db.session.commit()
        
        # 自动登录
        session['user_id'] = new_user.id
        session['username'] = new_user.username
        session['user_name'] = new_user.user_name
        return jsonify({'success': True, 'message': '注册成功'})
    logged_in = 'user_id' in session
    username = session.get('username', '')
    return render_template('register.html', logged_in=logged_in, username=username)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # 验证用户
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['user_name'] = user.user_name
            return jsonify({'success': True, 'message': '登录成功'})
        else:
            return jsonify({'error': '用户名或密码错误'}), 401
    logged_in = 'user_id' in session
    username = session.get('username', '')
    return render_template('login.html', logged_in=logged_in, username=username)

# 管理员密码（实际生产环境应该使用更安全的方式存储）
ADMIN_PASSWORD = 'HlYsXlT'

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin_reset_password', methods=['GET', 'POST'])
def admin_reset_password():
    if request.method == 'POST':
        admin_password = request.form.get('admin_password')
        username = request.form.get('username')
        
        # 验证管理员密码
        if admin_password != ADMIN_PASSWORD:
            return jsonify({'success': False, 'message': '管理员密码错误'}), 401
        
        # 查找用户
        user = User.query.filter_by(username=username).first()
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        # 重置密码为默认值
        default_password = '123456'
        user.password = generate_password_hash(default_password)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'密码重置成功！用户名：{username}，新密码：{default_password}'
        })
    
    return render_template('admin_reset_password.html')

@app.route('/delete_account')
def delete_account():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 删除用户的所有记录
    MovieRecord.query.filter_by(user_id=session['user_id']).delete()
    
    # 删除用户账号
    User.query.filter_by(id=session['user_id']).delete()
    
    # 提交数据库事务
    db.session.commit()
    
    # 清除会话
    session.clear()
    
    return redirect(url_for('index'))

@app.route('/records')
def records():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 获取类型参数
    record_type = request.args.get('type', '')
    
    # 根据类型过滤记录
    if record_type:
        user_records = MovieRecord.query.filter_by(user_id=session['user_id'], type=record_type).order_by(MovieRecord.watch_time.desc().nullslast()).all()
    else:
        user_records = MovieRecord.query.filter_by(user_id=session['user_id']).order_by(MovieRecord.watch_time.desc().nullslast()).all()
    
    return render_template('records.html', 
                           records=user_records, 
                           selected_type=record_type, 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           logged_in=True)

@app.route('/add_record', methods=['GET', 'POST'])
def add_record():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'GET':
        return render_template('add_record.html', 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']), 
                               logged_in=True)
    
    # 处理POST请求
    data = request.form
    
    # 处理观看时间
    watch_time = None
    if data.get('watch_time'):
        watch_time = datetime.strptime(data['watch_time'], '%Y-%m-%dT%H:%M')
    
    new_record = MovieRecord(
        user_id=session['user_id'],
        title=data['title'],
        type=data['type'],
        watch_time=watch_time,
        location=data.get('location'),
        people=data.get('people'),
        event=data.get('event'),
        rating=data.get('rating', type=int),
        comment=data.get('comment'),
        tags=data.get('tags')
    )
    
    db.session.add(new_record)
    db.session.commit()
    return redirect(url_for('records'))

@app.route('/record_detail/<int:record_id>')
def record_detail(record_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    record = MovieRecord.query.get(record_id)
    if not record or record.user_id != session['user_id']:
        return redirect(url_for('records'))
    
    return render_template('record_detail.html', 
                           record=record, 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           logged_in=True)

@app.route('/edit_record/<int:record_id>', methods=['GET', 'POST'])
def edit_record(record_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    record = MovieRecord.query.get(record_id)
    if not record or record.user_id != session['user_id']:
        return redirect(url_for('records'))
    
    if request.method == 'POST':
        title = request.form.get('title')
        record_type = request.form.get('type')
        watch_time_str = request.form.get('watch_time')
        location = request.form.get('location')
        people = request.form.get('people')
        event = request.form.get('event')
        rating = request.form.get('rating')
        comment = request.form.get('comment')
        tags = request.form.get('tags')
        
        # 处理标签，将逗号替换为顿号
        if tags:
            tags = tags.replace(',', '、')
        
        # 更新记录
        record.title = title
        record.type = record_type
        record.watch_time = datetime.datetime.fromisoformat(watch_time_str) if watch_time_str else None
        record.location = location
        record.people = people
        record.event = event
        record.rating = int(rating) if rating else None
        record.comment = comment
        record.tags = tags
        
        db.session.commit()
        return redirect(url_for('record_detail', record_id=record.id))
    
    user = User.query.get(session['user_id'])
    return render_template('edit_record.html', 
                           record=record, 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           logged_in=True)

@app.route('/delete_record/<int:record_id>')
def delete_record(record_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    record = MovieRecord.query.get(record_id)
    if record and record.user_id == session['user_id']:
        db.session.delete(record)
        db.session.commit()
    return redirect(url_for('records'))

@app.route('/search')
def search():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('search.html', 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           logged_in=True)

@app.route('/api/search', methods=['GET'])
def api_search():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    query = request.args.get('query', '')
    record_type = request.args.get('type', '')
    
    results = MovieRecord.query.filter_by(user_id=session['user_id'])
    if query:
        results = results.filter(MovieRecord.title.contains(query) | MovieRecord.comment.contains(query))
    if record_type:
        results = results.filter_by(type=record_type)
    
    records = results.order_by(MovieRecord.watch_time.desc()).all()
    
    return jsonify([{
        'id': r.id,
        'title': r.title,
        'type': r.type,
        'watch_time': r.watch_time.strftime('%Y-%m-%d %H:%M') if r.watch_time else None,
        'rating': r.rating,
        'comment': r.comment,
        'tags': r.tags
    } for r in records])

@app.route('/api/stats')
def api_stats():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 统计数据
    total_records = MovieRecord.query.filter_by(user_id=session['user_id']).count()
    type_stats = db.session.query(
        MovieRecord.type,
        db.func.count(MovieRecord.id)
    ).filter_by(user_id=session['user_id']).group_by(MovieRecord.type).all()
    
    return jsonify({
        'total_records': total_records,
        'type_stats': [{'type': t[0], 'count': t[1]} for t in type_stats]
    })

@app.route('/api/share/<int:record_id>', methods=['POST'])
def api_share(record_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    record = MovieRecord.query.get(record_id)
    if not record or record.user_id != session['user_id']:
        return jsonify({'error': '记录不存在或无权限'}), 404
    
    record.is_shared = True
    db.session.commit()
    return jsonify({'success': True, 'message': '分享成功'})

@app.route('/api/unshare/<int:record_id>', methods=['POST'])
def api_unshare(record_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    record = MovieRecord.query.get(record_id)
    if not record or record.user_id != session['user_id']:
        return jsonify({'error': '记录不存在或无权限'}), 404
    
    record.is_shared = False
    db.session.commit()
    return jsonify({'success': True, 'message': '取消分享成功'})

@app.route('/shared')
def shared_records():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 获取所有分享的记录（包括其他用户的）
    shared_records = MovieRecord.query.filter_by(is_shared=True).order_by(MovieRecord.created_at.desc()).all()
    return render_template('shared.html', 
                           records=shared_records, 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           logged_in=True)

# 个人主页路由
@app.route('/profile')
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('profile.html', 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           user_id=session['user_id'], 
                           logged_in=True)



# 修改用户名称路由
@app.route('/profile/update_user_name', methods=['POST'])
def update_user_name():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    new_user_name = request.form.get('new_user_name')
    if not new_user_name:
        return render_template('profile.html', 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']), 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='用户名称不能为空')
    
    # 更新用户名称
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))
    user.user_name = new_user_name
    session['user_name'] = new_user_name
    db.session.commit()
    
    return render_template('profile.html', 
                           username=session['username'], 
                           user_name=session['user_name'], 
                           user_id=session['user_id'], 
                           logged_in=True, 
                           success_message='用户名称修改成功')

# 修改密码路由
@app.route('/profile/update_password', methods=['POST'])
def update_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    # 验证输入
    if not current_password or not new_password or not confirm_password:
        return render_template('profile.html', 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']), 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='所有密码字段不能为空')
    
    # 验证新密码和确认密码是否一致
    if new_password != confirm_password:
        return render_template('profile.html', 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']), 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='新密码和确认密码不一致')
    
    # 验证当前密码是否正确
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))
    if not check_password_hash(user.password, current_password):
        return render_template('profile.html', 
                               username=session['username'], 
                               user_name=session.get('user_name', session['username']), 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='当前密码不正确')
    
    # 更新密码
    user.password = generate_password_hash(new_password)
    db.session.commit()
    
    return render_template('profile.html', 
                           username=session['username'], 
                           user_name=session.get('user_name', session['username']), 
                           user_id=session['user_id'], 
                           logged_in=True, 
                           success_message='密码修改成功')

# 导出数据路由
@app.route('/profile/export_data')
def export_data():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 获取用户的所有记录
    user_records = MovieRecord.query.filter_by(user_id=session['user_id']).all()
    
    # 构建导出数据
    export_data = {
        'username': session['username'],
        'user_name': session.get('user_name', session['username']),
        'user_id': session['user_id'],
        'export_time': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
        'records': [
            {
                'id': record.id,
                'title': record.title,
                'type': record.type,
                'watch_time': record.watch_time.strftime('%Y-%m-%d %H:%M') if record.watch_time else None,
                'location': record.location,
                'people': record.people,
                'event': record.event,
                'rating': record.rating,
                'comment': record.comment,
                'tags': record.tags,
                'is_shared': record.is_shared,
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for record in user_records
        ]
    }
    
    # 转换为JSON并下载
    import json
    from flask import send_file
    import io
    
    json_data = json.dumps(export_data, ensure_ascii=False, indent=2)
    buffer = io.BytesIO(json_data.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(buffer, 
                     mimetype='application/json', 
                     as_attachment=True, 
                     download_name=f'movie_history_{session["username"]}_{datetime.utcnow().strftime("%Y%m%d")}.json')

# 导出Excel数据路由
@app.route('/profile/export_data_excel')
def export_data_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 获取用户的所有记录
    user_records = MovieRecord.query.filter_by(user_id=session['user_id']).all()
    
    # 创建Excel文件
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "观影记录"
    
    # 设置表头
    headers = ['标题', '类型', '观看时间', '地点', '同行人员', '活动', '评分', '评论', '标签', '是否分享', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    for row, record in enumerate(user_records, 2):
        ws.cell(row=row, column=1, value=record.title)
        ws.cell(row=row, column=2, value=record.type)
        ws.cell(row=row, column=3, value=record.watch_time.strftime('%Y-%m-%d %H:%M') if record.watch_time else '')
        ws.cell(row=row, column=4, value=record.location or '')
        ws.cell(row=row, column=5, value=record.people or '')
        ws.cell(row=row, column=6, value=record.event or '')
        ws.cell(row=row, column=7, value=record.rating)
        ws.cell(row=row, column=8, value=record.comment or '')
        ws.cell(row=row, column=9, value=record.tags or '')
        ws.cell(row=row, column=10, value='是' if record.is_shared else '否')
        ws.cell(row=row, column=11, value=record.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, 
                     download_name=f'movie_history_{session["username"]}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx')

# 导入数据路由
@app.route('/profile/import_data', methods=['POST'])
def import_data():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if 'data_file' not in request.files:
        return render_template('profile.html', 
                               username=session['username'], 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='请选择要导入的文件')
    
    file = request.files['data_file']
    if file.filename == '':
        return render_template('profile.html', 
                               username=session['username'], 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message='请选择要导入的文件')
    
    try:
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'overwrite')
        print(f"Import mode: {import_mode}")
        print(f"Form data: {request.form}")
        
        # 如果是覆盖模式，清除用户现有记录
        if import_mode == 'overwrite':
            print("Overwriting existing records...")
            MovieRecord.query.filter_by(user_id=session['user_id']).delete()
        else:
            print("Adding to existing records...")
        
        if file.filename.endswith('.json'):
            # 处理JSON文件
            import json
            data = json.load(file)
            
            # 导入新记录
            for record_data in data.get('records', []):
                # 处理标签，确保使用顿号分隔
                tags = record_data.get('tags')
                if tags and ',' in tags:
                    # 如果标签中包含逗号，替换为顿号
                    tags = tags.replace(',', '、')
                
                new_record = MovieRecord(
                    user_id=session['user_id'],
                    title=record_data['title'],
                    type=record_data['type'],
                    watch_time=datetime.strptime(record_data['watch_time'], '%Y-%m-%d %H:%M'),
                    location=record_data.get('location'),
                    people=record_data.get('people'),
                    event=record_data.get('event'),
                    rating=record_data.get('rating'),
                    comment=record_data.get('comment'),
                    tags=tags,
                    is_shared=record_data.get('is_shared', False)
                )
                db.session.add(new_record)
        elif file.filename.endswith('.xlsx'):
            # 处理Excel文件
            from openpyxl import load_workbook
            import io
            
            # 读取Excel文件
            buffer = io.BytesIO(file.read())
            wb = load_workbook(buffer)
            ws = wb.active
            
            # 跳过表头，从第二行开始读取数据
            for row in range(2, ws.max_row + 1):
                title = ws.cell(row=row, column=1).value
                record_type = ws.cell(row=row, column=2).value
                watch_time_str = ws.cell(row=row, column=3).value
                location = ws.cell(row=row, column=4).value
                people = ws.cell(row=row, column=5).value
                event = ws.cell(row=row, column=6).value
                rating = ws.cell(row=row, column=7).value
                comment = ws.cell(row=row, column=8).value
                tags = ws.cell(row=row, column=9).value
                is_shared_str = ws.cell(row=row, column=10).value
                
                # 验证必填字段
                if not title or not record_type:
                    continue
                
                # 处理标签，确保使用顿号分隔
                if tags and isinstance(tags, str) and ',' in tags:
                    # 如果标签中包含逗号，替换为顿号
                    tags = tags.replace(',', '、')
                
                # 处理观看时间
                watch_time = None
                if watch_time_str:
                    try:
                        if isinstance(watch_time_str, str):
                            watch_time = datetime.strptime(watch_time_str, '%Y-%m-%d %H:%M')
                        else:
                            # 如果是datetime对象
                            watch_time = watch_time_str
                    except:
                        pass
                
                # 处理是否分享
                is_shared = False
                if is_shared_str in ['是', 'true', 'True', '1']:
                    is_shared = True
                
                # 创建新记录
                new_record = MovieRecord(
                    user_id=session['user_id'],
                    title=title,
                    type=record_type,
                    watch_time=watch_time,
                    location=location,
                    people=people,
                    event=event,
                    rating=rating,
                    comment=comment,
                    tags=tags,
                    is_shared=is_shared
                )
                db.session.add(new_record)
        else:
            return render_template('profile.html', 
                                   username=session['username'], 
                                   user_id=session['user_id'], 
                                   logged_in=True, 
                                   error_message='请选择JSON或Excel格式的文件')
        
        db.session.commit()
        
        return render_template('profile.html', 
                               username=session['username'], 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               success_message=f'数据导入成功，导入模式: {import_mode}')
    
    except Exception as e:
        return render_template('profile.html', 
                               username=session['username'], 
                               user_id=session['user_id'], 
                               logged_in=True, 
                               error_message=f'导入失败：{str(e)}')

if __name__ == '__main__':
    app.run(debug=True)